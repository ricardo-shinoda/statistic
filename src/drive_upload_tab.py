import openpyxl
import os
from googleapiclient.discovery import build
from google.oauth2 import service_account
from googleapiclient.http import MediaFileUpload

# Especificação dos arquivos de origem e destino
source_file_path = '/home/ricardo/code/statistic/src/credit_card/xlsx/2023-09.xlsx'
target_file_path = '/home/ricardo/code/statistic/src/Controle.xlsx'

# Nome da aba de origem e nome da nova aba de destino
source_sheet_name = 'Sheet1'
new_target_sheet_name = 'new'

# Carrega a aba de origem usando openpyxl
source_workbook = openpyxl.load_workbook(source_file_path, data_only=True)
source_sheet = source_workbook[source_sheet_name]

# Carrega a planilha de destino (ou cria uma nova se não existir)
if os.path.exists(target_file_path):
    target_workbook = openpyxl.load_workbook(target_file_path)
    target_sheet = target_workbook[new_target_sheet_name]
else:
    target_workbook = openpyxl.Workbook()
    target_sheet = target_workbook.active
    target_sheet.title = new_target_sheet_name

# Copia os dados da aba de origem para a aba de destino, ignorando linhas em branco
for row in source_sheet.iter_rows(values_only=True):
    if any(row):
        target_sheet.append(row)

# Salva a planilha de destino
target_workbook.save(target_file_path)

# Faz o upload da planilha para o Google Drive

file_name = "2023-08"  # Nome do arquivo no Google Drive

SCOPES = ['https://www.googleapis.com/auth/drive']
SERVICE_ACCOUNT_FILE = 'service.account.json'
PARENT_FOLDER_ID = "1qyMZb6P7H5oaq2zxoaqs17VJWSnHNgnG"


def authenticate():
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    return creds


def upload_file(file_path):
    creds = authenticate()
    service = build('drive', 'v3', credentials=creds)

    file_metadata = {
        'name': f'{file_name}.xlsx',
        'parents': [PARENT_FOLDER_ID]
    }

    media_body = MediaFileUpload(file_path, resumable=True)

    file = service.files().create(
        body=file_metadata,
        media_body=media_body
    ).execute()


upload_file(target_file_path)
