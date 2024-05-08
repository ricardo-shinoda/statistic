import os
import shutil
import pandas as pd
import json
import openpyxl
from googleapiclient.http import MediaFileUpload
from google.oauth2 import service_account
from googleapiclient.discovery import build

source_directory = "/home/ricardo/Downloads/invoice"
target_extension = ".csv"
destination_directory = "/home/ricardo/code/statistic/src"
month = "2023-01"

# Move CSV files from source to destination directory
for root, _, files in os.walk(source_directory):
    for current_file in files:
        if current_file.endswith(target_extension):
            source_file_path = os.path.join(root, current_file)
            destination_file_name = 'invoice.csv'
            destination_file_path = os.path.join(
                destination_directory, destination_file_name)
            shutil.move(source_file_path, destination_file_path)
            break

# Load CSV file
df = pd.read_csv(os.path.join(destination_directory,
                 'invoice.csv'), delimiter=';')

# Drop unnecessary rows
df = df[df['Descrição'] != 'Inclusao de Pagamento']

# Load description mapping from JSON file
with open('/home/ricardo/code/statistic/src/description.json') as f:
    description_mapping = {item['original']: item['new']
                           for item in json.load(f)}

# Map categories
df['Mapped_Categoria'] = df['Descrição'].map(description_mapping)
df['Categoria'] = df['Mapped_Categoria'].combine_first(df['Categoria'])
df = df.drop(columns=['Mapped_Categoria'])

# Remove leading and trailing whitespaces from column names
df.columns = df.columns.str.strip()

# Group by category and user
try:
    category_sum = df.groupby('Categoria')['Valor (em R$)'].sum().reset_index()
    user_sum = df.groupby('Nome no Cartão')[
        'Valor (em R$)'].sum().reset_index()
    total_expenses = user_sum['Valor (em R$)'].sum()
except KeyError as e:
    print(f"KeyError: {
          e}. Check if the column 'Valor (em R$)' is present in the DataFrame.")

# Save data to Excel
with pd.ExcelWriter(f'/home/ricardo/code/statistic/src/credit_card/xlsx/{month}.xlsx', engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Data', index=False, startrow=0, startcol=0)
    category_sum.to_excel(writer, sheet_name='Data', index=False,
                          startrow=0, startcol=df.shape[1] + 1, header=True)
    user_sum.to_excel(writer, sheet_name='Data', index=False,
                      startrow=0, startcol=df.shape[1] + 3, header=True)
    total_expenses_df = pd.DataFrame(
        {'Nome no Cartão': ['Total Expenses'], 'Valor (em R$)': [total_expenses]})
    total_expenses_df.to_excel(writer, sheet_name='Data', index=False,
                               startrow=user_sum.shape[0] + 2, startcol=df.shape[1] + 3, header=True)

# Save data to JSON
with open(f'/home/ricardo/code/statistic/src/credit_card/json/{month}.json', 'w', encoding='utf-8') as json_file:
    json.dump(df.to_dict(orient='records'),
              json_file, indent=4, ensure_ascii=False)

# Save data to Controle.xlsx
source_file_path = f'/home/ricardo/code/statistic/src/credit_card/xlsx/{
    month}.xlsx'
target_file_path = '/home/ricardo/code/statistic/src/Controle.xlsx'
source_sheet_name = 'Data'
new_target_sheet_name = f'{month}'

source_workbook = openpyxl.load_workbook(source_file_path, data_only=True)
source_sheet = source_workbook[source_sheet_name]

if os.path.exists(target_file_path):
    target_workbook = openpyxl.load_workbook(target_file_path)
else:
    target_workbook = openpyxl.Workbook()

if new_target_sheet_name in target_workbook.sheetnames:
    target_sheet = target_workbook[new_target_sheet_name]
else:
    target_sheet = target_workbook.create_sheet(title=new_target_sheet_name)

for row in source_sheet.iter_rows(values_only=True):
    target_sheet.append(row)

target_workbook.save(target_file_path)

# Upload Controle.xlsx to Google Drive
SCOPES = ['https://www.googleapis.com/auth/drive']
SERVICE_ACCOUNT_FILE = 'service.account.json'
PARENT_FOLDER_ID = "1qyMZb6P7H5oaq2zxoaqs17VJWSnHNgnG"


def authenticate():
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    return creds


def upload_file(file_path):
    creds = authenticate()
    service = build('drive', 'v3', credentials=creds)

    file_metadata = {'name': f'Controle_{
        month}.xlsx', 'parents': [PARENT_FOLDER_ID]}
    media_body = MediaFileUpload(file_path, resumable=True)

    file = service.files().create(body=file_metadata, media_body=media_body).execute()


upload_file(target_file_path)
